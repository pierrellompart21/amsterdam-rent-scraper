"""Export listings to Excel format."""

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from rich.console import Console

from amsterdam_rent_scraper.config.settings import EXCEL_FILENAME
from amsterdam_rent_scraper.models.listing import RentalListing

console = Console()


# Column order and display names
COLUMN_ORDER = [
    ("source_site", "Source"),
    ("title", "Title"),
    ("price_eur", "Price (EUR)"),
    ("address", "Address"),
    ("city", "City"),
    ("postal_code", "Postal Code"),
    ("surface_m2", "Area (m²)"),
    ("rooms", "Rooms"),
    ("bedrooms", "Bedrooms"),
    ("bathrooms", "Bathrooms"),
    ("furnished", "Furnished"),
    ("available_date", "Available"),
    ("deposit_eur", "Deposit (EUR)"),
    ("energy_label", "Energy"),
    ("pets_allowed", "Pets"),
    ("distance_km", "Distance (km)"),
    ("commute_time_bike_min", "Bike (min)"),
    ("commute_time_transit_min", "Transit (min)"),
    ("description_summary", "Summary"),
    ("pros", "Pros"),
    ("cons", "Cons"),
    ("agency", "Agency"),
    ("listing_url", "URL"),
    ("scraped_at", "Scraped At"),
]


def listings_to_dataframe(listings: list[dict | RentalListing]) -> pd.DataFrame:
    """Convert listings to a pandas DataFrame."""
    data = []
    for listing in listings:
        if isinstance(listing, RentalListing):
            row = listing.model_dump()
        else:
            row = listing
        data.append(row)

    df = pd.DataFrame(data)

    # Reorder columns based on COLUMN_ORDER
    ordered_cols = []
    for col_key, _ in COLUMN_ORDER:
        if col_key in df.columns:
            ordered_cols.append(col_key)

    # Add any remaining columns not in COLUMN_ORDER
    for col in df.columns:
        if col not in ordered_cols:
            ordered_cols.append(col)

    df = df[ordered_cols]
    return df


def style_excel_workbook(wb: Workbook) -> None:
    """Apply styling to the Excel workbook."""
    ws = wb.active

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Border style
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Style header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Style data rows
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"


def export_to_excel(
    listings: list[dict | RentalListing], output_dir: Path, filename: str = None
) -> Path:
    """Export listings to an Excel file."""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    filename = filename or EXCEL_FILENAME
    filepath = output_dir / filename

    console.print(f"[cyan]Exporting {len(listings)} listings to Excel...[/]")

    df = listings_to_dataframe(listings)

    # Rename columns to display names
    column_map = {key: display for key, display in COLUMN_ORDER}
    df = df.rename(columns=column_map)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Rental Listings"

    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx + 1, column=c_idx, value=value)

    # Apply styling
    style_excel_workbook(wb)

    # Save
    wb.save(filepath)
    console.print(f"[green]Excel file saved: {filepath}[/]")

    return filepath
